"""
Public-records demand-signal engine.

Third leg of the acquisition stack. operator_outreach.py recruits SUPPLY,
b2b_outreach.py sells recurring COMMERCIAL demand — this module intercepts
one-off demand at the moment it is created, from Florida public records
(Ch. 119): code-enforcement citations (owner has a compliance deadline to
remove junk), probate filings (estate cleanouts), eviction filings (unit
clearouts ~30-45 days out).

Pipeline (daily): pluggable per-jurisdiction fetchers -> normalize ->
dedupe on (record_type, jurisdiction, case_number) -> store DemandRecord ->
digest email + admin review (/api/admin/demand/signals).

v1 deliberately does NOT contact anyone. These are consumer records, not
business listings: no cold email (addresses are postal anyway), no cold SMS
(TCPA). Rows feed the digest; a human decides who gets a letter or a call.
Fetchers hit only public, no-login pages, identify themselves, and rate-limit.
"""

import datetime as dt
import io
import logging
import os
import re
import time

import requests

logger = logging.getLogger(__name__)

_UA = "UmuveRecordsBot/1.0 (+https://goumuve.com; public-records research)"
_TIMEOUT = 25

# Violation text that maps to junk-removal demand. Anything not matching one of
# these is dropped — we only want cited property owners who literally must
# remove items to comply.
_JUNK_VIOLATION_RE = re.compile(
    r"junk|debris|trash|garbage|rubbish|outdoor storage|outside storage|"
    r"abandoned (?:item|furniture|appliance|vehicle)|inoperable vehicle|"
    r"bulk|solid waste|unsightly|accumulation|litter|refuse|hoard",
    re.IGNORECASE,
)


def _get(url, **kw):
    kw.setdefault("timeout", _TIMEOUT)
    headers = kw.pop("headers", {})
    headers.setdefault("User-Agent", _UA)
    return requests.get(url, headers=headers, **kw)


def _parse_date(text):
    if not text:
        return None
    text = str(text).strip()[:40]
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y",
                "%Y-%m-%dT%H:%M:%S", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(text.split(".")[0], fmt)
        except ValueError:
            continue
    # epoch millis (ArcGIS)
    if text.isdigit() and len(text) == 13:
        try:
            return dt.datetime.utcfromtimestamp(int(text) / 1000.0)
        except (ValueError, OverflowError):
            pass
    return None


# --------------------------------------------------------------------------- #
# Fetchers — each returns a list of normalized dicts:
#   {record_type, jurisdiction, case_number, filed_date, property_address,
#    city, zip, party_name, party_address, details, source_url}
# Fetchers must never raise (return [] on failure) and must stay polite:
# public no-login endpoints only, one request at a time, sleep between pages.
# --------------------------------------------------------------------------- #

# Source notes (verified 2026-08-21):
# - PBC unincorporated: Special Magistrate agenda PDFs at
#   discover.pbc.gov/pzb/CodeCompliance/Hearing/{YYYY}/Agenda{MMDDYYYY}.PDF,
#   monthly (~first Wednesday). Richest source: respondent name + mailing
#   address + situs + verbatim violation text. ~96 cases/agenda, ~30
#   junk-relevant. Implemented below.
# - West Palm Beach eGovPlus (onestopshop.wpbgov.com): mechanically perfect
#   (plain POST + detail pages w/ owner mailing address) but the dataset is
#   FROZEN at ~July 2023 — verified via three broad owner-name samples. Do
#   not wire it up until the city's current system surfaces (same trap as
#   Fort Lauderdale's ArcGIS CodeCaseTracker, stale since 2019).
# - Phase 2 candidates: Broward BCS POSSE (date-range search), Click2Gov CE
#   (Lake Worth Beach + Royal Palm Beach, same platform), Legistar API
#   (Delray body "Code Enforcement Board" / Wellington body 193).
# - Probate + evictions are NOT fetched — the PBC Clerk portal prohibits
#   bots, and the Clerk sells weekly extracts (Decedent 07 / Evictions 06).
#   Those arrive as files and enter through ingest_clerk_report() below.


def _wednesdays(year, month):
    d = dt.date(year, month, 1)
    d += dt.timedelta(days=(2 - d.weekday()) % 7)  # first Wednesday
    out = []
    while d.month == month:
        out.append(d)
        d += dt.timedelta(days=7)
    return out


_PBC_AGENDA_URL = "https://discover.pbc.gov/pzb/CodeCompliance/Hearing/{y}/Agenda{m:02d}{d:02d}{y}.PDF"
_PBC_CASE_RE = re.compile(r"C-\d{4}-\d{8}")


def fetch_pbc_magistrate_agenda():
    """PBC Code Compliance Special Magistrate agendas -> junk-relevant cases.

    Tries every Wednesday of the current and previous month (hearing is
    usually the first Wednesday; agendas post ahead of it). Dedupe upstream
    makes re-parsing a known agenda harmless."""
    today = dt.date.today()
    months = [(today.year, today.month)]
    prev = (today.replace(day=1) - dt.timedelta(days=1))
    months.append((prev.year, prev.month))

    rows = []
    for year, month in months:
        for wed in _wednesdays(year, month):
            url = _PBC_AGENDA_URL.format(y=year, m=wed.month, d=wed.day)
            try:
                resp = _get(url)
            except requests.RequestException:
                continue
            time.sleep(1.0)
            if resp.status_code != 200 or not resp.content[:4] == b"%PDF":
                continue
            rows.extend(_parse_pbc_agenda(resp.content, url, hearing_date=wed))
            break  # one agenda per month
    return rows


def _parse_pbc_agenda(pdf_bytes, source_url, hearing_date):
    from pypdf import PdfReader

    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        full = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        logger.warning("pbc agenda pdf unreadable: %s", source_url)
        return []

    rows = []
    for block in re.split(r"(?=Respondent:)", full)[1:]:
        m = _PBC_CASE_RE.search(block)
        if not m or not _JUNK_VIOLATION_RE.search(block):
            continue
        resp_m = re.search(r"Respondent:\s*(?:CEO:)?([^\n]+)", block)
        mail_m = re.search(r"^([0-9][^\n]+,\s*FL\s*\d{5}(?:-\d{4})?)\s*$", block, re.M)
        situs_m = re.search(r"Case No:Situs Address:\s*C-\d{4}-\d{8}\s*([^\n]+)", block)
        issued_m = re.search(r"Issued:\s*(\d{2}/\d{2}/\d{4})", block)
        # Which cited code sections are junk-shaped (for the reviewer's eyes).
        cited = [c.strip() for c in re.findall(r"Code:([^\n]+)", block)
                 if re.search(r"14-35|6\.D\.1", c)]
        situs = (situs_m.group(1).strip() if situs_m else None)
        city = None
        if situs and "," in situs:
            parts = [p.strip() for p in situs.split(",")]
            if len(parts) >= 2:
                city = parts[-2] if parts[-1].upper().startswith("FL") else parts[-1]
        mail = mail_m.group(1).strip() if mail_m else None
        zip_m = re.search(r"(\d{5})(?:-\d{4})?\s*$", mail or "")
        rows.append({
            "record_type": "code_violation",
            "jurisdiction": "pbc_county",
            "case_number": m.group(0),
            "filed_date": _parse_date(issued_m.group(1)) if issued_m else None,
            "property_address": situs,
            "city": city,
            "zip": zip_m.group(1) if zip_m else None,
            "party_name": resp_m.group(1).strip()[:200] if resp_m else None,
            "party_address": mail,
            "details": "SM hearing {} | cited: {}".format(
                hearing_date.isoformat(), "; ".join(cited) or "junk/debris violation text matched"),
            "source_url": source_url,
        })
    return rows


# Registry — the cycle iterates this. Keep observe-only sources here; anything
# that would CONTACT a person does not belong in this module.
FETCHERS = {
    "pbc_magistrate": fetch_pbc_magistrate_agenda,
}


# --------------------------------------------------------------------------- #
# Store
# --------------------------------------------------------------------------- #
def _store(db, DemandRecord, rows):
    """Insert new rows, skipping (record_type, jurisdiction, case_number) dupes.
    Returns (inserted, list_of_new_signals)."""
    if not rows:
        return 0, []
    existing = {
        (r.record_type, r.jurisdiction, r.case_number)
        for r in db.session.query(
            DemandRecord.record_type, DemandRecord.jurisdiction, DemandRecord.case_number
        ).all()
    }
    inserted, new_signals = 0, []
    for row in rows:
        case = (row.get("case_number") or "").strip()[:80]
        if not case:
            continue
        key = (row["record_type"], row["jurisdiction"], case)
        if key in existing:
            continue
        existing.add(key)
        sig = DemandRecord(
            record_type=row["record_type"],
            jurisdiction=row["jurisdiction"],
            case_number=case,
            filed_date=row.get("filed_date"),
            property_address=(row.get("property_address") or "")[:300] or None,
            city=(row.get("city") or "")[:80] or None,
            zip=(row.get("zip") or "")[:10] or None,
            party_name=(row.get("party_name") or "")[:200] or None,
            party_address=(row.get("party_address") or "")[:300] or None,
            details=(row.get("details") or "")[:2000] or None,
            source_url=(row.get("source_url") or "")[:500] or None,
            status="new",
        )
        db.session.add(sig)
        new_signals.append(sig)
        inserted += 1
    db.session.commit()
    return inserted, new_signals


# --------------------------------------------------------------------------- #
# Digest
# --------------------------------------------------------------------------- #
def _digest_html(report, new_signals):
    rows_html = ""
    for s in new_signals[:40]:
        rows_html += (
            "<tr><td style='padding:4px 8px'>{}</td><td style='padding:4px 8px'>{}</td>"
            "<td style='padding:4px 8px'>{}</td><td style='padding:4px 8px'>{}</td></tr>"
        ).format(
            s.record_type.replace("_", " "),
            s.property_address or s.party_name or "—",
            s.city or "",
            (s.details or "")[:90],
        )
    per_source = "".join(
        "<li>{}: <b>{}</b> fetched, <b>{}</b> new{}</li>".format(
            name, r["fetched"], r["inserted"],
            " — <span style='color:#C52222'>ERROR: {}</span>".format(r["error"]) if r.get("error") else "",
        )
        for name, r in sorted(report["sources"].items())
    )
    return """\
<div style="font-family:Arial,sans-serif;font-size:14px;color:#1a1a1a">
  <h2>Demand signals — daily sweep</h2>
  <p><b>{new}</b> new signals ({cv} code-violation / {pb} probate / {ev} eviction).
     Review + mark: GET /api/admin/demand/signals · CSV: /api/admin/demand/signals.csv</p>
  <ul>{per_source}</ul>
  <table style="border-collapse:collapse;font-size:13px" border="1" cellspacing="0">
    <tr style="background:#f4f4f4"><th style="padding:4px 8px">Type</th>
        <th style="padding:4px 8px">Property / party</th>
        <th style="padding:4px 8px">City</th><th style="padding:4px 8px">Details</th></tr>
    {rows}
  </table>
  <p style="color:#777;font-size:12px">v1 is observe-only: no one is contacted automatically.
     Code-violation rows are filtered to junk/debris/outdoor-storage citations only.</p>
</div>""".format(
        new=report["inserted"],
        cv=report["by_type"].get("code_violation", 0),
        pb=report["by_type"].get("probate", 0),
        ev=report["by_type"].get("eviction", 0),
        per_source=per_source,
        rows=rows_html or "<tr><td colspan='4' style='padding:6px'>none today</td></tr>",
    )


# --------------------------------------------------------------------------- #
# Cycle
# --------------------------------------------------------------------------- #
def run_demand_records_cycle(app):
    """Daily entrypoint. Never raises — logs + returns a report dict."""
    with app.app_context():
        from models import db, DemandRecord
        from notifications import send_email

        report = {"inserted": 0, "by_type": {}, "sources": {}}
        all_new = []
        for name, fetcher in FETCHERS.items():
            src = {"fetched": 0, "inserted": 0, "error": None}
            try:
                rows = fetcher() or []
                src["fetched"] = len(rows)
                inserted, new_signals = _store(db, DemandRecord, rows)
                src["inserted"] = inserted
                all_new.extend(new_signals)
                for s in new_signals:
                    report["by_type"][s.record_type] = report["by_type"].get(s.record_type, 0) + 1
                report["inserted"] += inserted
            except Exception as exc:  # fetchers shouldn't raise, but belt+braces
                logger.exception("demand_records fetcher %s failed", name)
                db.session.rollback()
                src["error"] = str(exc)[:200]
            report["sources"][name] = src
            time.sleep(1.0)

        report_to = os.environ.get("OUTREACH_REPORT_TO", "") or os.environ.get("ADMIN_EMAIL", "")
        if report_to and (report["inserted"] or any(s.get("error") for s in report["sources"].values())):
            try:
                send_email(report_to, "Umuve demand signals — {} new".format(report["inserted"]),
                           _digest_html(report, all_new))
            except Exception:
                logger.warning("demand records digest email failed")

        logger.info("demand records sweep: %s", report)
        return report


# --------------------------------------------------------------------------- #
# Clerk-report ingest (probate / evictions)
#
# The PBC Clerk's eCaseView portal prohibits automated access, but the Clerk
# SELLS the exact extracts we need via Clerk Cart: "Decedent 07" (weekly
# probate filings incl. personal-rep + attorney addresses, formal/summary
# admin only) and "Evictions 06" (weekly eviction filings). Those arrive as
# Excel/CSV; this ingests them. Column layouts vary, so headers are matched
# fuzzily and unmapped columns are preserved in details.
# --------------------------------------------------------------------------- #

_HEADER_MAP = {
    "case_number": ("case number", "case no", "case_no", "casenumber", "case #", "ucn", "uniform case"),
    "party_name": ("personal representative", "personal rep", "petitioner", "plaintiff",
                   "party name", "pr name", "representative"),
    "decedent": ("decedent", "deceased"),
    "defendant": ("defendant", "respondent", "tenant"),
    # NOTE: order matters — _map_headers checks fields in dict order, and
    # party_address's bare "address" alias would swallow "property address"
    # if it were checked first. Most-specific field goes first.
    "property_address": ("property address", "premises", "location", "site address"),
    "party_address": ("address", "mailing address", "pr address", "representative address",
                      "plaintiff address", "party address"),
    "city": ("city",),
    "zip": ("zip", "zip code", "postal"),
    "filed_date": ("filing date", "filed", "file date", "date filed", "open date"),
    "case_type": ("case type", "type", "court type", "division"),
}


def _map_headers(header_row):
    """Map raw column headers -> canonical field names, fuzzily."""
    mapping = {}
    for idx, raw in enumerate(header_row):
        lo = str(raw or "").strip().lower()
        if not lo:
            continue
        for field, aliases in _HEADER_MAP.items():
            if field not in mapping.values() and any(a in lo for a in aliases):
                mapping[idx] = field
                break
    return mapping


def _report_rows(file_storage):
    """Yield rows (lists of cell values) from an uploaded CSV or XLSX."""
    filename = (file_storage.filename or "").lower()
    if filename.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        wb.close()
    else:  # csv / txt
        import csv as _csv

        text = file_storage.read().decode("utf-8-sig", errors="replace")
        for row in _csv.reader(io.StringIO(text)):
            yield row


def ingest_clerk_report(db, DemandRecord, file_storage, record_type, jurisdiction="pbc_clerk"):
    """Parse an uploaded Clerk Cart report into DemandRecord rows.

    record_type: 'probate' | 'eviction'. Returns a report dict; never raises
    on malformed rows (they're counted + skipped)."""
    assert record_type in ("probate", "eviction")
    mapping, parsed, skipped = None, [], 0
    for row in _report_rows(file_storage):
        if not any(str(c or "").strip() for c in row):
            continue
        if mapping is None:
            mapping = _map_headers(row)
            if "case_number" in mapping.values():
                continue  # was a real header row
            mapping = None  # preamble line — keep looking
            continue
        rec = {field: str(row[idx] or "").strip() for idx, field in mapping.items() if idx < len(row)}
        if not rec.get("case_number"):
            skipped += 1
            continue
        # Probate: the person to write to is the PR; decedent goes in details.
        # Eviction: plaintiff/landlord is the buyer; premises = where the junk is.
        details_bits = []
        if rec.get("case_type"):
            details_bits.append(rec["case_type"])
        if record_type == "probate" and rec.get("decedent"):
            details_bits.append("Estate of {}".format(rec["decedent"]))
        if record_type == "eviction" and rec.get("defendant"):
            details_bits.append("v. {}".format(rec["defendant"]))
        parsed.append({
            "record_type": record_type,
            "jurisdiction": jurisdiction,
            "case_number": rec["case_number"],
            "filed_date": _parse_date(rec.get("filed_date")),
            "property_address": rec.get("property_address"),
            "city": rec.get("city"),
            "zip": rec.get("zip"),
            "party_name": rec.get("party_name"),
            "party_address": rec.get("party_address"),
            "details": " | ".join(details_bits) or None,
            "source_url": None,
        })
    inserted, _ = _store(db, DemandRecord, parsed)
    return {
        "rows_parsed": len(parsed),
        "rows_skipped": skipped,
        "inserted": inserted,
        "duplicates": len(parsed) - inserted,
        "columns_mapped": sorted(set(mapping.values())) if mapping else [],
    }


def signals_csv(db, DemandRecord, status=None, days=None):
    """CSV export used by the admin endpoint."""
    import csv

    query = db.session.query(DemandRecord)
    if status:
        query = query.filter(DemandRecord.status == status)
    if days:
        cutoff = dt.datetime.utcnow() - dt.timedelta(days=days)
        query = query.filter(DemandRecord.created_at >= cutoff)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["type", "jurisdiction", "case_number", "filed_date", "property_address",
                "city", "zip", "party_name", "party_address", "details", "status", "source_url"])
    for s in query.order_by(DemandRecord.created_at.desc()).limit(2000).all():
        w.writerow([s.record_type, s.jurisdiction, s.case_number,
                    s.filed_date.date().isoformat() if s.filed_date else "",
                    s.property_address or "", s.city or "", s.zip or "",
                    s.party_name or "", s.party_address or "",
                    (s.details or "").replace("\n", " ")[:300], s.status, s.source_url or ""])
    return buf.getvalue()
